# import all our modules
import openpyxl
from os import path
from openpyxl import cell
# # Trying to program a login program

# Write different functions used in the code.


# check if workbook has been created, if so just load it do not make another one
def load_workbook(wb_path):
    if path.exists(wb_path):
        return openpyxl.load_workbook(wb_path)
    return openpyxl.Workbook()


# defining of variables

wb_path = "userdata.xlsx"
wb = load_workbook(wb_path)

# ws = wb["userdata"]
ws = wb.active


# Define home screen, the user is redirected to this in case of an error.
def home_screen():
    while True:
        print("Please select one of the options beneath.")
        print("(1): Create an account.""\n" "(2): Log in.""\n" "(3): Show Usernames.""\n" 
              "(4): Show Passwords.""\n" "(5): Exit.")
        print("\n")
        selection = int(input("Please enter your selection: "))
        if selection == 1:
            create_account()
        elif selection == 2:
            log_in()
        elif selection == 3:
            print(usernames)
        elif selection == 4:
            print(passwords)
        elif selection == 5:
            exit()


usernames = []
passwords = []


# Function to create an account.
def create_account():
    print("Hello, let's create an account.")
    print("\n")
    username = input("Please enter a username: ")
    password = input("Please enter a password: ")
    if username in usernames:
        print("This username is already taken")
    else:
        usernames.append(username)
    passwords.append(password)
    for username in usernames:
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value=username)
    for password in passwords:
        ws.cell(row=1, column=2, value=password)
    wb.save(wb_path)
    home_screen()


def log_in():
    print("Let's log in, please enter your data beneath""\n")
    a = ("\n" + input("Please enter your username: "))
    b  = ("\n" + input("Please enter your password: "))
    if ws.values([usernames]) == a and ws.values(passwords) == b:
        print("You have successfully logged in")
        home_screen()

    # for i in range(1, ws.max_row):
# if ws.cell(row=row, column=0).value == username:
# for j in range(i, ws.max_column):
# print(ws.cell(row=i, column=j).value)


# creating usernames and password lists.


# getting input from the user to put into the lists


home_screen()

#
#
#     userdata.close()
#
#
#
# if selection == 1:
#     create_account(username,password)
# elif selection == 2:
#     log_in(username, password)
# else:
#     Print("Please select one either (1) or (2).")



